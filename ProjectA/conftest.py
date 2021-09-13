import pytest
import openpyxl

book = openpyxl.load_workbook("C:/Users/nikos/OneDrive/Υπολογιστής/BootCamps/GitDemo/ProjectA/Sheets/UserSheet.xlsx")
sheet = book.active

def ListOfUsers(ind=None):
    sheet_list = []
    for i in range(2, sheet.max_row + 1):
        user_dict = {}
        for j in range(2,sheet.max_column + 1):
            user_dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
        sheet_list.append(user_dict)
    if ind != None:
        return sheet_list[ind]
    else:
        return sheet_list

print(ListOfUsers())

@pytest.fixture(params=ListOfUsers())
def user(request):
    return request.param

@pytest.fixture(params=["lot", "little"])
def quantity(request):
    return request.param

